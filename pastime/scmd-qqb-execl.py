# !/magedu/bin/env python
# coding=utf-8

from openpyxl import load_workbook, Workbook

def app(addr):
    gm_data = 1
    gm_r = 'A'
    gc_r = 'B'
    g = 2
    nc = 2
    bj_r = 'I'
    bj_c = 2
    while gm_data:
        g1 = str(g)
        nc1 = str(nc)
        gm_data = es[gm_r+g1].value
        gc_data = es[gc_r+g1].value
        print(u'修改%s' % gm_r+nc1)
        ns[gm_r+nc1] = addr[gm_data[0]]+gm_data
        print(u'修改%s' % gc_r+nc1)
        ns[gc_r+nc1] = gc_data
        for i in ['C','D','E']:
            bj_c1 = str(bj_c)
            bj_data = es[bj_r+bj_c1].value
            print(u'修改%s' % i+nc1)
            ns[i+nc1] = bj_data
            bj_c += 1
        nc += 1
        g += 3

if __name__ == "__main__":
    addr = {
        '0': 'sz',
        '1': 'sz',
        '2': 'sz',
        '3': 'sz',
        '4': 'sz',
        '5': 'sz',
        '6': 'sh',
        '7': 'sz',
        '8': 'sz',
        '9': 'sz'
    }
    new = Workbook()
    ns = new.active
    try:
        execl = load_workbook('c:\\1.xlsx')
        es = execl.active
    except:
        print(u'''执行错误，请检查
        1.表格是否为2007版本
        2.表格是否正在被打开使用
        3.命名是否正确（1.xlsx）''')
    ns['A1'] = '股票代码'
    ns['B1'] = '股票名称'
    ns['C1'] = '5d'
    ns['D1'] = '10d'
    ns['E1'] = '1m'
    try:
        app(addr)
    except:
        new.save('c:\\new.xlsx')
        print(u'新表保存完成，请检查数据')
    input(u"程序完成，请按任意键退出！")